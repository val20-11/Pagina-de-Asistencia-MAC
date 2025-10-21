#!/usr/bin/env python
"""
Script para cargar datos iniciales desde archivos Excel a la base de datos.
Este script importa:
- Eventos desde Conferencias MAC Agenda Completa.xlsx
- Estudiantes desde Student.xlsx
- Perfiles de Asistentes desde AssistantProfile.xlsx
- Estadísticas de Asistencia desde AttendanceStats.xlsx
"""

import os
import sys
import django
from pathlib import Path

# Configurar Django
BASE_DIR = Path(__file__).resolve().parent
sys.path.append(str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mac_attendance.settings')
django.setup()

import openpyxl
from datetime import datetime, time
from django.contrib.auth.models import User
from authentication.models import Student, AssistantProfile, Asistente, SystemConfiguration
from events.models import Event
from attendance.models import Attendance, AttendanceStats
from django.utils import timezone

def import_events(file_path):
    """Importar eventos desde Conferencias MAC Agenda Completa.xlsx"""
    print(f"\n📅 Importando eventos desde {file_path}...")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    events_created = 0
    events_updated = 0

    # Saltar el encabezado (primera fila)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Si no hay título, saltar
            continue

        title = row[0]
        speaker = row[1] if len(row) > 1 and row[1] else ""
        date_str = row[2] if len(row) > 2 and row[2] else None
        start_time_str = row[3] if len(row) > 3 and row[3] else None
        end_time_str = row[4] if len(row) > 4 and row[4] else None
        event_type = row[5] if len(row) > 5 and row[5] else "Conferencia"
        modality = row[6] if len(row) > 6 and row[6] else "Presencial"
        location = row[7] if len(row) > 7 and row[7] else ""
        description = row[8] if len(row) > 8 and row[8] else ""
        is_active = bool(row[9]) if len(row) > 9 and row[9] else True

        if not date_str or not start_time_str:
            print(f"  ⚠️  Saltando evento sin fecha u hora: {title}")
            continue

        # Convertir fecha
        if isinstance(date_str, datetime):
            event_date = date_str.date()
        elif isinstance(date_str, str):
            try:
                event_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                try:
                    event_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                except:
                    print(f"  ⚠️  Formato de fecha inválido para {title}: {date_str}")
                    continue
        else:
            event_date = date_str

        # Convertir hora de inicio
        if isinstance(start_time_str, time):
            start_time = start_time_str
        elif isinstance(start_time_str, str):
            try:
                start_time = datetime.strptime(start_time_str, '%H:%M:%S').time()
            except:
                try:
                    start_time = datetime.strptime(start_time_str, '%H:%M').time()
                except:
                    print(f"  ⚠️  Formato de hora inválido para {title}: {start_time_str}")
                    continue
        else:
            start_time = start_time_str

        # Convertir hora de fin
        end_time = None
        if end_time_str:
            if isinstance(end_time_str, time):
                end_time = end_time_str
            elif isinstance(end_time_str, str):
                try:
                    end_time = datetime.strptime(end_time_str, '%H:%M:%S').time()
                except:
                    try:
                        end_time = datetime.strptime(end_time_str, '%H:%M').time()
                    except:
                        pass

        # Crear o actualizar evento (sin validaciones para permitir fechas pasadas)
        try:
            event = Event.objects.get(title=title, date=event_date)
            # Actualizar usando update() para evitar validaciones
            Event.objects.filter(id=event.id).update(
                description=description,
                location=location,
                start_time=start_time,
                end_time=end_time,
                is_active=is_active
            )
            created = False
        except Event.DoesNotExist:
            # Crear sin validaciones usando el método directo de la base
            event = Event(
                title=title,
                date=event_date,
                description=description,
                location=location,
                start_time=start_time,
                end_time=end_time,
                is_active=is_active
            )
            # Guardar sin llamar a clean() para evitar validación de fechas pasadas
            super(Event, event).save(force_insert=True)
            created = True

        if created:
            events_created += 1
            print(f"  ✅ Evento creado: {title} - {event_date} {start_time}")
        else:
            events_updated += 1

    print(f"\n✅ Eventos importados: {events_created} creados, {events_updated} actualizados")
    return events_created + events_updated


def import_students(file_path):
    """Importar estudiantes desde Student.xlsx"""
    print(f"\n👨‍🎓 Importando estudiantes desde {file_path}...")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    students_created = 0
    students_updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Si no hay número de cuenta, saltar
            continue

        account_number = str(row[0]).strip()
        full_name = row[1] if len(row) > 1 and row[1] else "Sin nombre"

        student, created = Student.objects.update_or_create(
            account_number=account_number,
            defaults={
                'full_name': full_name,
                'user_type': 'student'
            }
        )

        if created:
            students_created += 1
            print(f"  ✅ Estudiante creado: {account_number} - {full_name}")
        else:
            students_updated += 1

    print(f"\n✅ Estudiantes importados: {students_created} creados, {students_updated} actualizados")
    return students_created + students_updated


def import_assistant_profiles(file_path):
    """Importar perfiles de asistentes desde AssistantProfile.xlsx"""
    print(f"\n👔 Importando perfiles de asistentes desde {file_path}...")

    from authentication.models import UserProfile
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    profiles_created = 0
    profiles_updated = 0
    profiles_skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Si no hay número de cuenta, saltar
            continue

        account_number = str(row[0]).strip()
        full_name = row[1] if len(row) > 1 and row[1] else "Sin nombre"

        # Verificar si ya existe un UserProfile con este account_number
        existing_profile = UserProfile.objects.filter(account_number=account_number).first()

        if existing_profile:
            # Si existe y es estudiante, actualizar a asistente
            if existing_profile.user_type == 'student':
                print(f"  ⚠️  Saltando {account_number} - ya existe como estudiante")
                profiles_skipped += 1
                continue
            else:
                # Si ya es asistente, actualizar
                existing_profile.full_name = full_name
                existing_profile.save()
                profiles_updated += 1
        else:
            # No existe, crear nuevo
            profile = AssistantProfile.objects.create(
                account_number=account_number,
                full_name=full_name,
                user_type='assistant'
            )
            profiles_created += 1
            print(f"  ✅ Perfil de asistente creado: {account_number} - {full_name}")

    print(f"\n✅ Perfiles de asistentes importados: {profiles_created} creados, {profiles_updated} actualizados, {profiles_skipped} saltados")
    return profiles_created + profiles_updated


def assign_permissions_to_assistant():
    """Asignar permisos al asistente para registrar asistencias"""
    print(f"\n🔑 Asignando permisos a asistentes...")

    from authentication.models import UserProfile
    # Obtener solo los UserProfile que son asistentes
    assistants = UserProfile.objects.filter(user_type='assistant')
    permissions_created = 0

    for assistant_profile in assistants:
        # Crear permiso de asistente (Asistente model)
        asistente, created = Asistente.objects.get_or_create(
            user_profile=assistant_profile,
            defaults={
                'can_manage_events': True
            }
        )

        if created:
            permissions_created += 1
            print(f"  ✅ Permisos asignados a: {assistant_profile.full_name}")

    print(f"\n✅ Permisos asignados: {permissions_created} asistentes")
    return permissions_created


def import_attendance_stats(file_path):
    """Importar estadísticas de asistencia desde AttendanceStats.xlsx"""
    print(f"\n📊 Importando estadísticas de asistencia desde {file_path}...")

    from authentication.models import UserProfile
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    stats_created = 0
    attendances_created = 0

    # Obtener eventos principales
    event1 = Event.objects.filter(title__icontains="Transformación Digital").first()
    event2 = Event.objects.filter(title__icontains="Finanzas").first()

    if not event1:
        print("  ⚠️  Evento 'Transformación Digital' no encontrado")
    if not event2:
        print("  ⚠️  Evento 'Finanzas' no encontrado")

    # Obtener el asistente para registered_by
    assistant = UserProfile.objects.filter(user_type='assistant').first()
    if not assistant:
        print("  ⚠️  No hay asistentes en el sistema. Creando asistente por defecto...")
        assistant = UserProfile.objects.create(
            account_number='99999999',
            full_name='Sistema - Importación Automática',
            user_type='assistant'
        )
        print(f"  ✅ Asistente creado: {assistant.full_name}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Si no hay número de cuenta, saltar
            continue

        account_number = str(row[0]).strip()
        full_name = row[1] if len(row) > 1 else ""
        attended_events = int(row[2]) if len(row) > 2 and row[2] else 0
        total_events = int(row[3]) if len(row) > 3 and row[3] else 18
        attendance_percentage = float(row[4]) if len(row) > 4 and row[4] else 0.0

        try:
            student = Student.objects.get(account_number=account_number)
        except Student.DoesNotExist:
            # Si el estudiante no existe, crearlo
            student = Student.objects.create(
                account_number=account_number,
                full_name=full_name,
                user_type='student'
            )
            print(f"  ℹ️  Estudiante creado desde stats: {account_number}")

        # Crear estadísticas
        stats, created = AttendanceStats.objects.update_or_create(
            student=student,
            defaults={
                'total_events': total_events,
                'attended_events': attended_events,
                'attendance_percentage': attendance_percentage
            }
        )

        if created:
            stats_created += 1

        # Crear asistencias según el número total
        if attended_events == 2:
            # Asistió a ambas conferencias
            if event1:
                try:
                    attendance = Attendance.objects.get(student=student, event=event1)
                    att_created = False
                except Attendance.DoesNotExist:
                    attendance = Attendance(
                        student=student,
                        event=event1,
                        timestamp=timezone.now(),
                        is_valid=True,
                        registered_by=assistant
                    )
                    # Guardar sin validaciones
                    super(Attendance, attendance).save(force_insert=True)
                    att_created = True
                    attendances_created += 1

            if event2:
                try:
                    attendance = Attendance.objects.get(student=student, event=event2)
                    att_created = False
                except Attendance.DoesNotExist:
                    attendance = Attendance(
                        student=student,
                        event=event2,
                        timestamp=timezone.now(),
                        is_valid=True,
                        registered_by=assistant
                    )
                    # Guardar sin validaciones
                    super(Attendance, attendance).save(force_insert=True)
                    att_created = True
                    attendances_created += 1

        elif attended_events == 1:
            # Asistió solo a "Transformación Digital"
            if event1:
                try:
                    attendance = Attendance.objects.get(student=student, event=event1)
                    att_created = False
                except Attendance.DoesNotExist:
                    attendance = Attendance(
                        student=student,
                        event=event1,
                        timestamp=timezone.now(),
                        is_valid=True,
                        registered_by=assistant
                    )
                    # Guardar sin validaciones
                    super(Attendance, attendance).save(force_insert=True)
                    att_created = True
                    attendances_created += 1

    print(f"\n✅ Estadísticas importadas: {stats_created} creadas")
    print(f"✅ Asistencias creadas: {attendances_created}")
    return stats_created


def configure_system_settings():
    """Configurar settings globales del sistema"""
    print(f"\n⚙️  Configurando sistema global...")

    config, created = SystemConfiguration.objects.get_or_create(
        id=1,
        defaults={
            'minimum_attendance_percentage': 80.0,
            'minutes_before_event': 10,  # 10 minutos antes
            'minutes_after_start': 25,   # 25 minutos después
        }
    )

    if created:
        print(f"  ✅ Configuración creada: 80% asistencia, 10min antes, 25min después")
    else:
        config.minimum_attendance_percentage = 80.0
        config.minutes_before_event = 10
        config.minutes_after_start = 25
        config.save()
        print(f"  🔄 Configuración actualizada")

    return config


def main():
    """Función principal"""
    print("="*60)
    print("🚀 CARGA DE DATOS INICIALES - SISTEMA DE ASISTENCIAS MAC")
    print("="*60)

    # Rutas a los archivos Excel
    base_path = Path(__file__).resolve().parent

    events_file = base_path / "Conferencias MAC Agenda Completa (1) (1).xlsx"
    students_file = base_path / "Student-2025-10-21.xlsx"
    assistants_file = base_path / "AssistantProfile-2025-10-21.xlsx"
    stats_file = base_path / "AttendanceStats-2025-10-21.xlsx"

    # Verificar que existan los archivos
    files_ok = True
    for file_path, name in [
        (events_file, "Eventos"),
        (students_file, "Estudiantes"),
        (assistants_file, "Asistentes"),
        (stats_file, "Estadísticas")
    ]:
        if not file_path.exists():
            print(f"❌ Archivo no encontrado: {name} - {file_path}")
            files_ok = False

    if not files_ok:
        print("\n⚠️  Algunos archivos no se encontraron. Abortando.")
        return

    # Importar datos
    try:
        # 1. Configurar sistema
        configure_system_settings()

        # 2. Importar eventos
        import_events(events_file)

        # 3. Importar estudiantes
        import_students(students_file)

        # 4. Importar asistentes
        import_assistant_profiles(assistants_file)

        # 5. Asignar permisos a asistentes
        assign_permissions_to_assistant()

        # 6. Importar estadísticas y asistencias
        import_attendance_stats(stats_file)

        print("\n" + "="*60)
        print("✅ IMPORTACIÓN COMPLETADA EXITOSAMENTE")
        print("="*60)

        # Mostrar resumen
        print(f"\n📊 Resumen de datos en la base de datos:")
        print(f"  - Eventos: {Event.objects.count()}")
        print(f"  - Estudiantes: {Student.objects.count()}")
        print(f"  - Asistentes: {AssistantProfile.objects.count()}")
        print(f"  - Permisos de asistentes: {Asistente.objects.count()}")
        print(f"  - Asistencias: {Attendance.objects.count()}")
        print(f"  - Estadísticas: {AttendanceStats.objects.count()}")

    except Exception as e:
        print(f"\n❌ Error durante la importación: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
