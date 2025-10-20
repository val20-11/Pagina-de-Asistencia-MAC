"""
Comando para importar asistencias desde un archivo Excel.

Formato esperado del Excel (asistencias.xlsx):
- Hoja: "Asistencias"
- Columnas:
  1. numero_cuenta (str): Número de cuenta del asistente (8 dígitos)
  2. nombre_completo (str): Nombre completo del asistente
  3. titulo_evento (str): Título del evento (debe existir)
  4. fecha_evento (str): Fecha del evento en formato YYYY-MM-DD o DD/MM/YYYY
  5. numero_cuenta_asistente (str): Número de cuenta del asistente que registra (8 dígitos)
  6. metodo_registro (str, opcional): manual, barcode, external (default: manual)
  7. notas (str, opcional): Notas adicionales

Uso:
    python manage.py import_attendance ruta/al/archivo.xlsx
"""

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from attendance.models import Attendance
from events.models import Event
from authentication.models import UserProfile, ExternalUser
import openpyxl
from datetime import datetime


class Command(BaseCommand):
    help = 'Importa asistencias desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta al archivo Excel')
        parser.add_argument(
            '--sheet',
            type=str,
            default='Asistencias',
            help='Nombre de la hoja de Excel (default: Asistencias)'
        )
        parser.add_argument(
            '--registrador',
            type=str,
            required=True,
            help='Número de cuenta del asistente que realiza la importación'
        )
        parser.add_argument(
            '--skip-validation',
            action='store_true',
            help='Omitir validación de horarios (para importar asistencias históricas)'
        )
        parser.add_argument(
            '--create-external',
            action='store_true',
            help='Crear automáticamente usuarios externos si no existen'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        sheet_name = options['sheet']
        registrador_cuenta = options['registrador']
        skip_validation = options['skip_validation']
        create_external = options['create_external']

        # Validar que el registrador exista y sea asistente
        try:
            registrador = UserProfile.objects.get(
                account_number=registrador_cuenta,
                user_type='assistant'
            )
        except UserProfile.DoesNotExist:
            raise CommandError(
                f'No se encontró un asistente con número de cuenta: {registrador_cuenta}'
            )

        try:
            workbook = openpyxl.load_workbook(excel_file)

            if sheet_name not in workbook.sheetnames:
                raise CommandError(f'La hoja "{sheet_name}" no existe en el archivo')

            sheet = workbook[sheet_name]

            # Validar encabezados
            expected_headers = [
                'numero_cuenta', 'nombre_completo', 'titulo_evento',
                'fecha_evento', 'metodo_registro', 'notas'
            ]

            headers = [cell.value.lower().strip() if cell.value else ''
                      for cell in sheet[1]]

            if not all(h in headers for h in expected_headers[:4]):
                self.stdout.write(
                    self.style.WARNING(
                        f'Encabezados esperados: {", ".join(expected_headers[:4])}'
                    )
                )
                raise CommandError(
                    'El archivo debe tener los encabezados correctos en la primera fila'
                )

            created_count = 0
            error_count = 0
            external_created = 0

            # Procesar filas (comenzando desde la fila 2)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraer datos
                    numero_cuenta = str(row[0]).strip() if row[0] else None
                    nombre_completo = row[1]
                    titulo_evento = row[2]
                    fecha_evento_str = row[3]
                    metodo_registro = row[4] if len(row) > 4 and row[4] else 'manual'
                    notas = row[5] if len(row) > 5 else None

                    # Validar datos requeridos
                    if not all([numero_cuenta, nombre_completo, titulo_evento, fecha_evento_str]):
                        self.stdout.write(
                            self.style.WARNING(
                                f'Fila {row_idx}: Datos incompletos, omitiendo...'
                            )
                        )
                        error_count += 1
                        continue

                    # Parsear fecha del evento
                    if isinstance(fecha_evento_str, datetime):
                        fecha_evento = fecha_evento_str.date()
                    else:
                        try:
                            fecha_evento = datetime.strptime(str(fecha_evento_str), '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                fecha_evento = datetime.strptime(str(fecha_evento_str), '%d/%m/%Y').date()
                            except ValueError:
                                self.stdout.write(
                                    self.style.WARNING(
                                        f'Fila {row_idx}: Formato de fecha inválido ({fecha_evento_str})'
                                    )
                                )
                                error_count += 1
                                continue

                    # Buscar el evento
                    try:
                        event = Event.objects.get(
                            title=titulo_evento,
                            date=fecha_evento
                        )
                    except Event.DoesNotExist:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Fila {row_idx}: Evento "{titulo_evento}" del {fecha_evento} no encontrado'
                            )
                        )
                        error_count += 1
                        continue
                    except Event.MultipleObjectsReturned:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Fila {row_idx}: Múltiples eventos con título "{titulo_evento}" del {fecha_evento}"'
                            )
                        )
                        error_count += 1
                        continue

                    # Buscar al estudiante o usuario externo
                    student = None
                    external_user = None

                    # Primero buscar en estudiantes regulares
                    try:
                        student = UserProfile.objects.get(
                            account_number=numero_cuenta,
                            user_type='student'
                        )
                    except UserProfile.DoesNotExist:
                        # Buscar en usuarios externos
                        try:
                            external_user = ExternalUser.objects.get(
                                account_number=numero_cuenta
                            )
                        except ExternalUser.DoesNotExist:
                            # Si create_external está activo, crear usuario externo
                            if create_external:
                                external_user = ExternalUser.objects.create(
                                    account_number=numero_cuenta,
                                    full_name=nombre_completo,
                                    status='approved',
                                    approved_by=registrador
                                )
                                external_created += 1
                                self.stdout.write(
                                    self.style.SUCCESS(
                                        f'  ↳ Usuario externo creado: {nombre_completo} ({numero_cuenta})'
                                    )
                                )
                            else:
                                self.stdout.write(
                                    self.style.WARNING(
                                        f'Fila {row_idx}: Usuario {numero_cuenta} no encontrado (usa --create-external para crear)'
                                    )
                                )
                                error_count += 1
                                continue

                    # Verificar si ya existe la asistencia
                    if student:
                        existing = Attendance.objects.filter(
                            student=student,
                            event=event,
                            is_valid=True
                        ).exists()
                    else:
                        existing = Attendance.objects.filter(
                            external_user=external_user,
                            event=event,
                            is_valid=True
                        ).exists()

                    if existing:
                        self.stdout.write(
                            self.style.WARNING(
                                f'Fila {row_idx}: Asistencia ya existe para {nombre_completo} en "{titulo_evento}"'
                            )
                        )
                        error_count += 1
                        continue

                    # Crear asistencia
                    if skip_validation:
                        # Crear directamente sin validación
                        attendance = Attendance(
                            student=student,
                            external_user=external_user,
                            event=event,
                            registered_by=registrador,
                            registration_method=metodo_registro,
                            notes=notas,
                            is_valid=True
                        )
                        # Guardar sin validación
                        super(Attendance, attendance).save()
                    else:
                        attendance = Attendance.objects.create(
                            student=student,
                            external_user=external_user,
                            event=event,
                            registered_by=registrador,
                            registration_method=metodo_registro,
                            notes=notas,
                            is_valid=True
                        )

                    created_count += 1
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'✓ Fila {row_idx}: Asistencia de "{nombre_completo}" a "{titulo_evento}" registrada'
                        )
                    )

                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f'✗ Fila {row_idx}: Error - {str(e)}'
                        )
                    )
                    error_count += 1

            # Resumen
            self.stdout.write('\n' + '='*60)
            self.stdout.write(self.style.SUCCESS(f'Asistencias creadas: {created_count}'))
            if external_created > 0:
                self.stdout.write(self.style.SUCCESS(f'Usuarios externos creados: {external_created}'))
            if error_count > 0:
                self.stdout.write(self.style.WARNING(f'Errores: {error_count}'))
            self.stdout.write('='*60)

        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {excel_file}')
        except Exception as e:
            raise CommandError(f'Error al procesar el archivo: {str(e)}')
