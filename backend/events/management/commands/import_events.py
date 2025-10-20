"""
Comando para importar eventos desde un archivo Excel.

Formato esperado del Excel (eventos.xlsx):
- Hoja: "Eventos"
- Columnas:
  1. titulo (str): Título de la ponencia
  2. descripcion (str): Descripción del evento
  3. tipo (str): conference, workshop, panel, seminar
  4. modalidad (str): presencial, online, hybrid
  5. ponente (str): Nombre del ponente
  6. fecha (str): Formato YYYY-MM-DD o DD/MM/YYYY
  7. hora_inicio (str): Formato HH:MM
  8. hora_fin (str): Formato HH:MM
  9. ubicacion (str): Ubicación o plataforma
  10. capacidad (int): Capacidad máxima
  11. enlace_reunion (str, opcional): URL para eventos online/hybrid
  12. id_reunion (str, opcional): ID de la reunión

Uso:
    python manage.py import_events ruta/al/archivo.xlsx
"""

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from events.models import Event
import openpyxl
from datetime import datetime


class Command(BaseCommand):
    help = 'Importa eventos desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta al archivo Excel')
        parser.add_argument(
            '--sheet',
            type=str,
            default='Eventos',
            help='Nombre de la hoja de Excel (default: Eventos)'
        )
        parser.add_argument(
            '--skip-validation',
            action='store_true',
            help='Omitir validación de fechas pasadas (para importar eventos históricos)'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        sheet_name = options['sheet']
        skip_validation = options['skip_validation']

        try:
            workbook = openpyxl.load_workbook(excel_file)

            if sheet_name not in workbook.sheetnames:
                raise CommandError(f'La hoja "{sheet_name}" no existe en el archivo')

            sheet = workbook[sheet_name]

            # Validar encabezados
            expected_headers = [
                'titulo', 'descripcion', 'tipo', 'modalidad', 'ponente',
                'fecha', 'hora_inicio', 'hora_fin', 'ubicacion', 'capacidad',
                'enlace_reunion', 'id_reunion'
            ]

            headers = [cell.value.lower().strip() if cell.value else ''
                      for cell in sheet[1]]

            if not all(h in headers for h in expected_headers[:10]):
                self.stdout.write(
                    self.style.WARNING(
                        f'Encabezados esperados: {", ".join(expected_headers[:10])}'
                    )
                )
                raise CommandError(
                    'El archivo debe tener los encabezados correctos en la primera fila'
                )

            created_count = 0
            error_count = 0

            # Procesar filas (comenzando desde la fila 2)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraer datos
                    titulo = row[0]
                    descripcion = row[1]
                    tipo = row[2]
                    modalidad = row[3]
                    ponente = row[4]
                    fecha_str = row[5]
                    hora_inicio_str = row[6]
                    hora_fin_str = row[7]
                    ubicacion = row[8]
                    capacidad = row[9]
                    enlace_reunion = row[10] if len(row) > 10 else None
                    id_reunion = row[11] if len(row) > 11 else None

                    # Validar datos requeridos
                    if not all([titulo, descripcion, tipo, modalidad, ponente,
                               fecha_str, hora_inicio_str, hora_fin_str, ubicacion]):
                        self.stdout.write(
                            self.style.WARNING(
                                f'Fila {row_idx}: Datos incompletos, omitiendo...'
                            )
                        )
                        error_count += 1
                        continue

                    # Parsear fecha
                    if isinstance(fecha_str, datetime):
                        fecha = fecha_str.date()
                    else:
                        try:
                            fecha = datetime.strptime(str(fecha_str), '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                fecha = datetime.strptime(str(fecha_str), '%d/%m/%Y').date()
                            except ValueError:
                                self.stdout.write(
                                    self.style.WARNING(
                                        f'Fila {row_idx}: Formato de fecha inválido ({fecha_str})'
                                    )
                                )
                                error_count += 1
                                continue

                    # Parsear horas
                    if isinstance(hora_inicio_str, datetime):
                        hora_inicio = hora_inicio_str.time()
                    else:
                        try:
                            hora_inicio = datetime.strptime(str(hora_inicio_str), '%H:%M').time()
                        except ValueError:
                            self.stdout.write(
                                self.style.WARNING(
                                    f'Fila {row_idx}: Formato de hora de inicio inválido ({hora_inicio_str})'
                                )
                            )
                            error_count += 1
                            continue

                    if isinstance(hora_fin_str, datetime):
                        hora_fin = hora_fin_str.time()
                    else:
                        try:
                            hora_fin = datetime.strptime(str(hora_fin_str), '%H:%M').time()
                        except ValueError:
                            self.stdout.write(
                                self.style.WARNING(
                                    f'Fila {row_idx}: Formato de hora de fin inválido ({hora_fin_str})'
                                )
                            )
                            error_count += 1
                            continue

                    # Crear evento (sin validación si skip_validation está activo)
                    if skip_validation:
                        # Crear directamente sin validación usando el método save del modelo base
                        event = Event(
                            title=titulo,
                            description=descripcion,
                            event_type=tipo,
                            modality=modalidad,
                            speaker=ponente,
                            date=fecha,
                            start_time=hora_inicio,
                            end_time=hora_fin,
                            location=ubicacion,
                            max_capacity=int(capacidad) if capacidad else 100,
                            meeting_link=enlace_reunion if enlace_reunion else None,
                            meeting_id=id_reunion if id_reunion else None,
                            is_active=True
                        )
                        # Guardar sin llamar a clean() - usar skip_validation
                        event.save(skip_validation=True)
                    else:
                        event = Event.objects.create(
                            title=titulo,
                            description=descripcion,
                            event_type=tipo,
                            modality=modalidad,
                            speaker=ponente,
                            date=fecha,
                            start_time=hora_inicio,
                            end_time=hora_fin,
                            location=ubicacion,
                            max_capacity=int(capacidad) if capacidad else 100,
                            meeting_link=enlace_reunion if enlace_reunion else None,
                            meeting_id=id_reunion if id_reunion else None,
                            is_active=True
                        )

                    created_count += 1
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'✓ Fila {row_idx}: Evento "{titulo}" creado exitosamente'
                        )
                    )

                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f'✗ Fila {row_idx}: Error - {str(e)}'
                        )
                    )
                    error_count += 1

            # Resumen
            self.stdout.write('\n' + '='*60)
            self.stdout.write(self.style.SUCCESS(f'Eventos creados: {created_count}'))
            if error_count > 0:
                self.stdout.write(self.style.WARNING(f'Errores: {error_count}'))
            self.stdout.write('='*60)

        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {excel_file}')
        except Exception as e:
            raise CommandError(f'Error al procesar el archivo: {str(e)}')
