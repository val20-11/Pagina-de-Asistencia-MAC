#!/usr/bin/env python
"""
Script para crear reporte de asistencia a conferencias específicas.
"""
import os
import sys
import django
import pandas as pd

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mac_attendance.settings.local')
django.setup()

from events.models import Event
from attendance.models import Attendance
from authentication.models import UserProfile

# Leer las matrículas del Excel
excel_path = '/app/Cálculo.xlsx'
df_matriculas = pd.read_excel(excel_path)
matriculas = df_matriculas.iloc[:, 0].tolist()  # Primera columna

print(f"Matrículas a consultar: {len(matriculas)}")
print(f"Matrículas: {matriculas[:10]}...")  # Primeras 10

# Buscar las conferencias específicas
conferencia1_titulo = "App Kachi México y su cultura"
conferencia2_titulo = "De principiante a protector: Luchando contra el Phishing"

# Buscar eventos que contengan estos títulos
eventos = Event.objects.all()
print(f"\nTotal de eventos en BD: {eventos.count()}")

# Buscar por título
conf1 = Event.objects.filter(title__icontains="App Kachi").first()
conf2 = Event.objects.filter(title__icontains="principiante a protector").first()

if conf1:
    print(f"\nConferencia 1 encontrada:")
    print(f"  ID: {conf1.id}")
    print(f"  Título: {conf1.title}")
    print(f"  Fecha: {conf1.date}")
    print(f"  Hora: {conf1.start_time}")
    print(f"  Ubicación: {conf1.location}")
else:
    print(f"\n¡ADVERTENCIA! No se encontró la conferencia '{conferencia1_titulo}'")
    print("Buscando conferencias similares...")
    similares = Event.objects.filter(title__icontains="Kachi")
    for e in similares:
        print(f"  - {e.title}")

if conf2:
    print(f"\nConferencia 2 encontrada:")
    print(f"  ID: {conf2.id}")
    print(f"  Título: {conf2.title}")
    print(f"  Fecha: {conf2.date}")
    print(f"  Hora: {conf2.start_time}")
    print(f"  Ubicación: {conf2.location}")
else:
    print(f"\n¡ADVERTENCIA! No se encontró la conferencia '{conferencia2_titulo}'")
    print("Buscando conferencias similares...")
    similares = Event.objects.filter(title__icontains="protector")
    for e in similares:
        print(f"  - {e.title}")

# Si no se encontraron las conferencias, mostrar todas
if not conf1 or not conf2:
    print("\n=== Listado de TODOS los eventos en la base de datos ===")
    for evento in Event.objects.all().order_by('-date'):
        print(f"  [{evento.id}] {evento.title} - {evento.date} {evento.start_time}")
    print(f"\nTotal: {Event.objects.count()} eventos")

# Procesar asistencias si se encontraron ambas conferencias
if conf1 and conf2:
    print("\n=== Procesando asistencias ===")

    # Crear listas para el reporte
    reporte_data = []

    for matricula in matriculas:
        # Buscar el estudiante
        # Las matrículas del Excel tienen 9 dígitos, las de la BD tienen 8
        # Necesitamos eliminar el último dígito
        matricula_bd = str(matricula)[:-1]

        try:
            estudiante = UserProfile.objects.get(account_number=matricula_bd)

            # Verificar asistencia a cada conferencia
            asistio_conf1 = Attendance.objects.filter(
                student=estudiante,
                event=conf1,
                is_valid=True
            ).exists()

            asistio_conf2 = Attendance.objects.filter(
                student=estudiante,
                event=conf2,
                is_valid=True
            ).exists()

            # Determinar a cuántas asistió
            if asistio_conf1 and asistio_conf2:
                asistencia_status = "Ambas conferencias"
            elif asistio_conf1:
                asistencia_status = "Solo App Kachi"
            elif asistio_conf2:
                asistencia_status = "Solo Phishing"
            else:
                asistencia_status = "Ninguna"

            reporte_data.append({
                'Matrícula': matricula,
                'Nombre': estudiante.full_name,
                'Asistencia': asistencia_status,
                'App Kachi México y su cultura': '✓' if asistio_conf1 else '✗',
                'De principiante a protector (Phishing)': '✓' if asistio_conf2 else '✗'
            })

        except UserProfile.DoesNotExist:
            reporte_data.append({
                'Matrícula': matricula,
                'Nombre': 'No encontrado en BD',
                'Asistencia': 'N/A',
                'App Kachi México y su cultura': 'N/A',
                'De principiante a protector (Phishing)': 'N/A'
            })

    # Crear DataFrame y guardarlo
    df_reporte = pd.DataFrame(reporte_data)

    # Ordenar: primero los que fueron a ambas, luego a una, luego ninguna
    orden_asistencia = {
        'Ambas conferencias': 1,
        'Solo App Kachi': 2,
        'Solo Phishing': 3,
        'Ninguna': 4,
        'N/A': 5
    }
    df_reporte['_orden'] = df_reporte['Asistencia'].map(orden_asistencia)
    df_reporte = df_reporte.sort_values('_orden').drop('_orden', axis=1)

    # Guardar en Excel
    output_path = '/app/Reporte_Asistencia_Conferencias.xlsx'

    # Crear un Excel con formato mejorado
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_reporte.to_excel(writer, sheet_name='Reporte de Asistencia', index=False)

        # Obtener el workbook y worksheet
        workbook = writer.book
        worksheet = writer.sheets['Reporte de Asistencia']

        # Ajustar ancho de columnas
        worksheet.column_dimensions['A'].width = 15  # Matrícula
        worksheet.column_dimensions['B'].width = 35  # Nombre
        worksheet.column_dimensions['C'].width = 25  # Asistencia
        worksheet.column_dimensions['D'].width = 30  # Conf 1
        worksheet.column_dimensions['E'].width = 35  # Conf 2

        # Aplicar formato a la cabecera
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Aplicar formato a las celdas de datos
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # Centrar las columnas de checkmarks
            for i in [0, 3, 4]:  # Matrícula y checkmarks
                row[i].alignment = Alignment(horizontal='center')

            # Color según asistencia
            asistencia = row[2].value
            if asistencia == 'Ambas conferencias':
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                row[2].fill = fill
                row[2].font = Font(bold=True, color='006100')
            elif 'Solo' in str(asistencia):
                fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                row[2].fill = fill
                row[2].font = Font(color='9C6500')
            elif asistencia == 'Ninguna':
                fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                row[2].fill = fill
                row[2].font = Font(color='9C0006')

    print(f"\n✓ Reporte creado exitosamente en: {output_path}")
    print(f"\nResumen:")
    print(f"  - Total de estudiantes: {len(reporte_data)}")
    print(f"  - Asistieron a ambas: {len([r for r in reporte_data if r['Asistencia'] == 'Ambas conferencias'])}")
    print(f"  - Solo App Kachi: {len([r for r in reporte_data if r['Asistencia'] == 'Solo App Kachi'])}")
    print(f"  - Solo Phishing: {len([r for r in reporte_data if r['Asistencia'] == 'Solo Phishing'])}")
    print(f"  - Ninguna: {len([r for r in reporte_data if r['Asistencia'] == 'Ninguna'])}")
    print(f"  - No encontrados en BD: {len([r for r in reporte_data if r['Asistencia'] == 'N/A'])}")
else:
    print("\n¡ERROR! No se pudieron encontrar las conferencias especificadas.")
    print("Por favor verifica los títulos de las conferencias.")
